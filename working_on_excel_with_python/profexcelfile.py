import openpyxl as xl  # give a name to whole tag

from openpyxl.chart import BarChart , Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename) # loading file
    sheet= wb["Sheet1"]

    for i in range(2, sheet.max_row+1):
        cell=sheet.cell(i,3)
        corrected_price=cell.value*0.9
        corrected_price_cell =sheet.cell(i,4)
        corrected_price_cell.value=corrected_price
    
    values =Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart =BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,"e2")

    wb.save(filename)
        