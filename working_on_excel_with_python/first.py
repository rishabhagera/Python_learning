import openpyxl as xl  # give a name to whole tag

from openpyxl.chart import BarChart , Reference

wb = xl.load_workbook("C:/Users/DELL/Desktop/python_course/working_on_excel_with_python/transaction.xlsx") # loading file

sheet= wb["Sheet1"]
cell = sheet["a1"]
#cell1 = sheet.cell(1,1)
#print(cell.value)

# total_row=(sheet.max_row)
# print(total_row)

# for i in range(1,sheet.max_row + 1):
#     print(i)

for i in range(2, sheet.max_row+1):
    cell=sheet.cell(i,3)
    #print(cell.value)
    corrected_price=cell.value*0.9
    #print(corrected_price)
    corrected_price_cell =sheet.cell(i,4)
    corrected_price_cell.value=corrected_price
  
  
values =Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart =BarChart()
chart.add_data(values)
sheet.add_chart(chart,"e2")


wb.save("C:/Users/DELL/Desktop/python_course/working_on_excel_with_python/transaction2.xlsx")
    