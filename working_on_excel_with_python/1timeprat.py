#pip install openpyxl

import openpyxl as xl

from openpyxl.chart import BarChart , Reference

wb = xl.load_workbook("C:/Users/DELL/Desktop/python_course/working_on_excel_with_python/transaction.xlsx")
sheet = wb["Sheet1"]

cell=sheet["a1"]
print(cell.value)
#print(wb)

for i in range(1,sheet.max_row+1):
    print(i)
    
for i in range(2,sheet.max_row+1):
    cell=sheet.cell(i,3)
    print(cell.value)
    new_price=cell.value*0.9
    correct_price=sheet.cell(i,4)
    correct_price.value=new_price

val=Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4
          )
chart=BarChart()
chart.add_data(val)
sheet.add_chart(chart,'e2')
    
wb.save("C:/Users/DELL/Desktop/python_course/working_on_excel_with_python/transaction4.xlsx")